"""
Excel 템플릿 생성기
"""

from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ..schemas import CashSchema, PositionSchema, TransactionSchema
from ..utils.logger import logger


class TemplateGenerator:
    """Excel 템플릿 생성기"""

    def __init__(self):
        self.schemas = {
            'cash': CashSchema,
            'position': PositionSchema,
            'transaction': TransactionSchema
        }

    def create_template(self, date: str, input_dir: str) -> Dict[str, Any]:
        """
        Excel 템플릿을 생성합니다.

        Args:
            date: 날짜 (YYYY-MM-DD 형식)
            input_dir: 입력 디렉토리 경로

        Returns:
            생성 결과 딕셔너리
        """
        try:
            # 날짜 검증
            datetime.strptime(date, "%Y-%m-%d")

            # 입력 디렉토리 생성
            input_path = Path(input_dir)
            date_dir = input_path / date
            date_dir.mkdir(parents=True, exist_ok=True)

            # 템플릿 파일 경로
            template_path = date_dir / "manual.xlsx"

            # Excel 워크북 생성
            wb = openpyxl.Workbook()

            # 기본 시트 제거
            wb.remove(wb.active)

            # 각 스키마별로 시트 생성
            sheet_info = {}
            for sheet_name, schema_class in self.schemas.items():
                self._create_sheet(wb, sheet_name, schema_class)
                sheet_info[sheet_name] = len(schema_class.__dataclass_fields__)

            # 워크북 저장
            wb.save(template_path)

            logger.info(f"템플릿 생성 완료: {template_path}")

            return {
                'status': 'success',
                'file_path': str(template_path),
                'sheets': sheet_info
            }

        except ValueError:
            error_msg = f"날짜 형식 오류: {date} (YYYY-MM-DD 형식으로 입력하세요)"
            logger.error(error_msg)
            return {'status': 'error', 'message': error_msg}

        except Exception as e:
            error_msg = f"템플릿 생성 실패: {e}"
            logger.error(error_msg)
            return {'status': 'error', 'message': error_msg}

    def _create_sheet(self, wb: openpyxl.Workbook, sheet_name: str, schema_class) -> None:
        """개별 시트를 생성합니다."""

        # 시트 생성
        ws = wb.create_sheet(title=sheet_name)

        # 스키마 필드 정보 가져오기
        fields = schema_class.__dataclass_fields__

        # 헤더 행 생성
        headers = []
        for field_name, field_info in fields.items():
            # 필수/선택 필드 표시
            is_required = field_info.default is field_info.default_factory
            required_mark = " (필수)" if is_required else " (선택)"
            headers.append(f"{field_name}{required_mark}")

        # 헤더 스타일링
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # 헤더 행 추가
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 컬럼 너비 자동 조정
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

        # 예시 데이터 행 추가 (2-4행)
        self._add_example_data(ws, schema_class, headers)

        # 테두리 추가
        self._add_borders(ws, len(headers), 4)

    def _add_example_data(self, ws: openpyxl.worksheet.worksheet.Worksheet, schema_class, headers: List[str]):
        """예시 데이터를 추가합니다."""

        # 스키마별 예시 데이터
        examples = {
            'cash': [
                ["2024-01-01", "예금", "국민은행", "1000000", "KRW", "manual", "2024-01-01T10:00:00"],
                ["2024-01-01", "적금", "신한은행", "500000", "KRW", "manual", "2024-01-01T10:00:00"],
                ["2024-01-01", "현금", "지갑", "50000", "KRW", "manual", "2024-01-01T10:00:00"]
            ],
            'position': [
                ["2024-01-01", "증권사", "삼성전자", "005930", "10", "70000", "KRW", "manual", "2024-01-01T10:00:00"],
                ["2024-01-01", "증권사", "SK하이닉스", "000660", "5", "120000", "KRW", "manual", "2024-01-01T10:00:00"],
                ["2024-01-01", "증권사", "카카오", "035720", "2", "50000", "KRW", "manual", "2024-01-01T10:00:00"]
            ],
            'transaction': [
                ["2024-01-01", "국민은행", "입금", "1000000", "급여", "월급", "KRW", "1월 급여", "manual", "2024-01-01T10:00:00"],
                ["2024-01-01", "신한은행", "출금", "500000", "이체", "적금이체", "KRW", "적금 납입", "manual", "2024-01-01T10:00:00"],
                ["2024-01-01", "증권사", "출금", "100000", "투자", "주식매수", "KRW", "삼성전자 매수", "manual", "2024-01-01T10:00:00"]
            ]
        }

        sheet_name = schema_class.__name__.replace('Schema', '').lower()
        example_data = examples.get(sheet_name, [])

        # 예시 데이터 추가
        for row_idx, row_data in enumerate(example_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 예시 데이터 스타일링
        example_font = Font(color="808080", italic=True)
        for row in range(2, len(example_data) + 2):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).font = example_font

    def _add_borders(self, ws: openpyxl.worksheet.worksheet.Worksheet, num_cols: int, num_rows: int):
        """테두리를 추가합니다."""

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in range(1, num_rows + 1):
            for col in range(1, num_cols + 1):
                ws.cell(row=row, column=col).border = thin_border
