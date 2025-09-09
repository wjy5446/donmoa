"""
뱅크샐러드 Provider
"""

from datetime import datetime
from pathlib import Path
from typing import Dict, Optional, Any, List

import pandas as pd
import openpyxl

from ..schemas import CashSchema, PositionSchema, TransactionSchema
from ..utils.logger import logger
from .base import BaseProvider


class BanksaladProvider(BaseProvider):
    """뱅크샐러드 Excel 파일 파싱 Provider"""

    def __init__(self, name: str = "banksalad_csv", config: Optional[Dict[str, Any]] = None):
        super().__init__(name, config)

    def get_supported_extensions(self) -> List[str]:
        """지원하는 파일 확장자 목록을 반환합니다."""
        return ["xlsx"]

    def parse_raw(self, file_path: Path) -> Dict[str, pd.DataFrame]:
        """원본 데이터를 파싱합니다."""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            dict_datas = {
                "financial_status": pd.DataFrame(),
                "expenses_records": pd.DataFrame(),
            }

            #########################################################
            # 재무현황 데이터 파싱
            #########################################################
            if "뱅샐현황" not in workbook.sheetnames:
                logger.error("뱅샐현황 시트를 찾을 수 없습니다")
                return dict_datas

            sheet = workbook["뱅샐현황"]

            # "3.재무현황" 헤더 찾기
            header_cell = None
            for cell in sheet['B']:
                if cell.value and "3.재무현황" in str(cell.value):
                    header_cell = cell
                    break

            if not header_cell:
                logger.error("3.재무현황 헤더를 찾을 수 없습니다")
                return dict_datas

            # 재무현황 테이블 파싱
            table_start_row = header_cell.row + 3

            header_row = sheet[table_start_row]
            header = [header_row[1].value, header_row[2].value, header_row[4].value]

            datas = []
            tmp_type = None
            for row in sheet.iter_rows(min_row=table_start_row + 1, values_only=True):
                if not any(cell is not None for cell in row):
                    break
                if row[1] is not None:
                    tmp_type = row[1]
                if row[2] is None:
                    continue
                datas.append([tmp_type, row[2], row[4]])

            dict_datas["financial_status"] = pd.DataFrame(datas, columns=header)

            #########################################################
            # 가계부 내역 데이터 파싱
            #########################################################
            dict_datas["expenses_records"] = pd.read_excel(file_path, sheet_name="가계부 내역")

            logger.info(f"재무현황 데이터 파싱 완료: {len(dict_datas['financial_status'])}건")
            logger.info(f"가계부 내역 데이터 파싱 완료: {len(dict_datas['expenses_records'])}건")

            return dict_datas

        except Exception as e:
            logger.error(f"데이터 파싱 실패: {e}")
            return {
                "financial_status": pd.DataFrame(),
                "expenses_records": pd.DataFrame()
            }

    def parse_cash(self, data: Dict[str, pd.DataFrame]) -> List[CashSchema]:
        """현금 데이터를 파싱합니다"""
        df_financial_status = data["financial_status"]

        cash_datas = []
        for _, row in df_financial_status.iterrows():
            category_type = row["항목"]
            account_name = row["상품명"]
            balance = row["금액"]

            if int(balance) == 0:
                continue

            if not (
                "자유입출금" in category_type or
                "저축성" in category_type or
                "전자금융" in category_type or
                "현금" in category_type
            ):
                continue

            if "저축성" in category_type and "청약" not in account_name:
                continue

            cash_datas.append(CashSchema(
                date=datetime.now().strftime("%Y-%m-%d"),
                category=self._get_category_type(category_type),
                account=account_name,
                balance=balance,
                currency="KRW",
                provider=self.name,
                collected_at=datetime.now().isoformat(),
            ))

        return cash_datas

    def parse_positions(self, data: Dict[str, pd.DataFrame]) -> List[PositionSchema]:
        """포지션 데이터를 파싱합니다"""
        return []

    def parse_transactions(self, data: Dict[str, pd.DataFrame]) -> List[TransactionSchema]:
        """거래내역 데이터를 파싱합니다"""
        df_expenses_records = data["expenses_records"]

        # 최근 1개월 데이터만 필터링
        if "날짜" in df_expenses_records.columns:
            df_expenses_records["날짜"] = pd.to_datetime(df_expenses_records["날짜"])
            max_date = df_expenses_records["날짜"].max()
            start_date = max_date - pd.DateOffset(months=1)
            df_expenses_records = df_expenses_records[df_expenses_records["날짜"] >= start_date]

        transactions_datas = []
        for _, row in df_expenses_records.iterrows():
            date = row["날짜"].strftime("%Y-%m-%d") + "T" + row["시간"].strftime("%H:%M:%S")
            account = row["결제수단"]
            transaction_type = row["타입"]
            amount = row["금액"]
            category = row["대분류"]
            category_detail = row["소분류"]
            currency = "KRW"
            note = row["메모"]

            transactions_datas.append(TransactionSchema(
                date=date,
                account=account,
                transaction_type=transaction_type,
                amount=amount,
                category=category,
                category_detail=category_detail,
                currency=currency,
                note=note,
                provider=self.name,
                collected_at=datetime.now().isoformat(),
            ))

        return transactions_datas

    def _get_category_type(self, category_text: str) -> str:
        """카테고리 텍스트를 타입으로 변환합니다"""
        if "자유입출금" in category_text or "저축성" in category_text:
            return "은행"
        elif "전자금융" in category_text:
            return "페이"
        else:
            return "기타"
