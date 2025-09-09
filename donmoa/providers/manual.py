"""
수동 입력 데이터 Provider
"""

import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional, Any
import openpyxl

from .base import BaseProvider
from ..core.schemas import CashSchema, PositionSchema, TransactionSchema
from ..utils.logger import logger


class ManualProvider(BaseProvider):
    """수동 입력 데이터 Provider"""

    def __init__(self, name: str = "manual", config: Optional[Dict[str, Any]] = None):
        super().__init__(name, config)

    def get_supported_extensions(self) -> List[str]:
        """지원하는 파일 확장자 목록을 반환합니다."""
        return ["xlsx"]

    def parse_raw(self, file_path: Path) -> Dict[str, pd.DataFrame]:
        """Excel 파일을 파싱합니다."""
        try:
            # Excel 파일 읽기
            wb = openpyxl.load_workbook(file_path, data_only=True)

            data = {}
            sheet_mapping = {
                'position': 'position',
                'cash': 'cash',
                'transaction': 'transaction'
            }

            for sheet_name, data_key in sheet_mapping.items():
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]

                    # 데이터를 DataFrame으로 변환
                    rows = []
                    for row in ws.iter_rows(values_only=True):
                        if any(cell is not None for cell in row):  # 빈 행 제외
                            rows.append(row)

                    if rows:
                        # 첫 번째 행을 헤더로 사용 (필수/선택 표시 제거)
                        headers = [str(header).split(' (')[0] for header in rows[0]]
                        data_rows = rows[1:]  # 헤더 제외

                        # 빈 행 제거
                        data_rows = [row for row in data_rows if any(cell is not None for cell in row)]

                        if data_rows:
                            df = pd.DataFrame(data_rows, columns=headers)
                            data[data_key] = df
                        else:
                            data[data_key] = pd.DataFrame()
                    else:
                        data[data_key] = pd.DataFrame()
                else:
                    data[data_key] = pd.DataFrame()

            wb.close()
            return data

        except Exception as e:
            logger.error(f"Excel 파일 파싱 실패: {e}")
            return {
                'position': pd.DataFrame(),
                'cash': pd.DataFrame(),
                'transaction': pd.DataFrame()
            }

    def parse_cash(self, data: Dict[str, pd.DataFrame]) -> List[CashSchema]:
        """현금 데이터를 파싱합니다."""
        cash_list = []

        if 'cash' not in data or data['cash'].empty:
            return cash_list

        df = data['cash']

        for _, row in df.iterrows():
            try:
                # 필수 필드 검증
                required_fields = ['date', 'category', 'account', 'balance']
                if not all(field in df.columns for field in required_fields):
                    logger.warning("현금 데이터에 필수 필드가 없습니다")
                    continue

                # 데이터 변환
                cash = CashSchema(
                    date=self._format_date(row.get('date', '')),
                    category=str(row.get('category', '')),
                    account=str(row.get('account', '')),
                    balance=self._convert_to_number(row.get('balance', 0)),
                    currency=str(row.get('currency', 'KRW')),
                    provider=self.name,
                    collected_at=self._get_current_timestamp()
                )

                cash_list.append(cash)

            except Exception as e:
                logger.warning(f"현금 데이터 파싱 실패: {e}")
                continue

        return cash_list

    def parse_positions(self, data: Dict[str, pd.DataFrame]) -> List[PositionSchema]:
        """포지션 데이터를 파싱합니다."""
        position_list = []

        if 'position' not in data or data['position'].empty:
            return position_list

        df = data['position']

        for _, row in df.iterrows():
            try:
                # 필수 필드 검증
                required_fields = ['date', 'account', 'name', 'ticker', 'quantity', 'average_price']
                if not all(field in df.columns for field in required_fields):
                    logger.warning("포지션 데이터에 필수 필드가 없습니다")
                    continue

                # 데이터 변환
                position = PositionSchema(
                    date=self._format_date(row.get('date', '')),
                    account=str(row.get('account', '')),
                    name=str(row.get('name', '')),
                    ticker=str(row.get('ticker', '')),
                    quantity=self._convert_to_number(row.get('quantity', 0)),
                    average_price=self._convert_to_number(row.get('average_price', 0)),
                    currency=str(row.get('currency', 'KRW')),
                    provider=self.name,
                    collected_at=self._get_current_timestamp()
                )

                position_list.append(position)

            except Exception as e:
                logger.warning(f"포지션 데이터 파싱 실패: {e}")
                continue

        return position_list

    def parse_transactions(self, data: Dict[str, pd.DataFrame]) -> List[TransactionSchema]:
        """거래 데이터를 파싱합니다."""
        transaction_list = []

        if 'transaction' not in data or data['transaction'].empty:
            return transaction_list

        df = data['transaction']

        for _, row in df.iterrows():
            try:
                # 필수 필드 검증
                required_fields = ['date', 'account', 'transaction_type', 'amount', 'category']
                if not all(field in df.columns for field in required_fields):
                    logger.warning("거래 데이터에 필수 필드가 없습니다")
                    continue

                # 데이터 변환
                transaction = TransactionSchema(
                    date=self._format_date(row.get('date', '')),
                    account=str(row.get('account', '')),
                    transaction_type=str(row.get('transaction_type', '')),
                    amount=self._convert_to_number(row.get('amount', 0)),
                    category=str(row.get('category', '')),
                    category_detail=(str(row.get('category_detail', ''))
                                     if pd.notna(row.get('category_detail')) else None),
                    currency=str(row.get('currency', 'KRW')),
                    note=(str(row.get('note', ''))
                          if pd.notna(row.get('note')) else None),
                    provider=self.name,
                    collected_at=self._get_current_timestamp()
                )

                transaction_list.append(transaction)

            except Exception as e:
                logger.warning(f"거래 데이터 파싱 실패: {e}")
                continue

        return transaction_list

    def _find_input_file(self, input_dir: Path) -> Optional[Path]:
        """수동 입력 파일을 찾습니다."""
        if not input_dir.exists():
            return None

        # manual.xlsx 파일 찾기
        manual_file = input_dir / "manual.xlsx"
        if manual_file.exists():
            return manual_file

        return None

    def _get_current_timestamp(self) -> str:
        """현재 타임스탬프를 반환합니다."""
        from datetime import datetime
        return datetime.now().isoformat()
